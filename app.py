from fastapi import FastAPI
from fastapi.responses import HTMLResponse, JSONResponse
from pathlib import Path
from datetime import datetime
import csv
import html
import io
import urllib.request

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

app = FastAPI(title="Klimaat Dashboard Kas4")

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
EXCEL_FILE = DATA_DIR / "Overzicht_alle_dagen.xlsx"
CSV_DIR = DATA_DIR / "per_dag"
CSV_FILE = CSV_DIR / "Overzicht_alle_dagen.csv"
REMOTE_CSV_URL = "https://raw.githubusercontent.com/peekbv-rgb/historischvergelijk-uden/main/data/per_dag/Overzicht_alle_dagen.csv"

REFERENCE_KAS = 4
KASSEN = [1, 2, 3, 4, 5, 6]
KAS_COLORS = {
    1: "#d9480f",
    2: "#e67700",
    3: "#f08c00",
    4: "#1864ab",
    5: "#2b8a3e",
    6: "#087f5b",
}


def to_float(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", ".")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def fmt(value, digits=1):
    number = to_float(value)
    return "-" if number is None else f"{number:.{digits}f}"


def avg(values):
    nums = [v for v in values if v is not None]
    return sum(nums) / len(nums) if nums else None


def max_or_none(values):
    nums = [v for v in values if v is not None]
    return max(nums) if nums else None


def read_excel_rows():
    if load_workbook is None or not EXCEL_FILE.exists():
        return []
    workbook = load_workbook(EXCEL_FILE, read_only=True, data_only=True)
    sheet_name = "Alle dagen" if "Alle dagen" in workbook.sheetnames else workbook.sheetnames[0]
    worksheet = workbook[sheet_name]
    raw_rows = list(worksheet.iter_rows(values_only=True))
    if not raw_rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in raw_rows[0]]
    rows = []
    for raw in raw_rows[1:]:
        row = {key: value for key, value in zip(headers, raw) if key}
        if any(value is not None for value in row.values()):
            rows.append(row)
    return rows


def read_csv_text(text):
    return [dict(row) for row in csv.DictReader(io.StringIO(text))]


def read_local_csv_rows():
    files = []
    if CSV_FILE.exists():
        files = [CSV_FILE]
    elif CSV_DIR.exists():
        files = sorted(CSV_DIR.rglob("*.csv"))

    rows = []
    for path in files:
        try:
            rows.extend(read_csv_text(path.read_text(encoding="utf-8-sig")))
        except Exception:
            continue
    return rows


def read_remote_csv_rows():
    try:
        with urllib.request.urlopen(REMOTE_CSV_URL, timeout=10) as response:
            text = response.read().decode("utf-8-sig")
        return read_csv_text(text)
    except Exception:
        return []


def load_rows():
    rows = read_excel_rows()
    if rows:
        return rows, "Excel: data/Overzicht_alle_dagen.xlsx"

    rows = read_local_csv_rows()
    if rows:
        return rows, "CSV: data/per_dag/Overzicht_alle_dagen.csv"

    rows = read_remote_csv_rows()
    if rows:
        return rows, "GitHub CSV fallback"

    return [], "geen databestand gevonden"


def debug_info():
    local_csvs = []
    if CSV_DIR.exists():
        local_csvs = [str(p.relative_to(BASE_DIR)) for p in sorted(CSV_DIR.rglob("*.csv"))]
    rows, source = load_rows()
    return {
        "rows_count": len(rows),
        "source": source,
        "base_dir": str(BASE_DIR),
        "excel_exists": EXCEL_FILE.exists(),
        "csv_dir_exists": CSV_DIR.exists(),
        "csv_file_exists": CSV_FILE.exists(),
        "local_csvs": local_csvs,
        "remote_csv_url": REMOTE_CSV_URL,
    }


def get_date(row):
    for key in ("Datum", "datum", "Date", "date"):
        value = row.get(key)
        if value is None:
            continue
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")
        return str(value)[:10]
    return "-"


def parse_date(row):
    try:
        return datetime.strptime(get_date(row), "%Y-%m-%d")
    except Exception:
        return datetime.min


def day_label(date_text):
    if len(date_text) >= 10:
        return f"{date_text[8:10]}-{date_text[5:7]}"
    return date_text


def get_first(row, keys):
    for key in keys:
        if key in row:
            return to_float(row.get(key))
    return None


def get_tmax(row, kas):
    return get_first(row, [
        f"Temp_Afd{kas}_max",
        f"Temp_Kas{kas}_max",
        f"Tmax_Afd{kas}",
        f"Tmax_Kas{kas}",
        f"Temp_max_Afd{kas}",
        f"Temp_max_Kas{kas}",
    ])


def get_vdmax(row, kas):
    return get_first(row, [
        f"VD_Afd{kas}_max",
        f"VD_Kas{kas}_max",
        f"VDmax_Afd{kas}",
        f"VDmax_Kas{kas}",
        f"VD_max_Afd{kas}",
        f"VD_max_Kas{kas}",
    ])


def get_vdgem(row, kas):
    return get_first(row, [
        f"VD_Afd{kas}_gem",
        f"VD_Kas{kas}_gem",
        f"VD_Afd{kas}_avg",
        f"VD_Kas{kas}_avg",
        f"VD_gem_Afd{kas}",
        f"VD_gem_Kas{kas}",
    ])


def latest_rows(rows, limit=20):
    valid = [row for row in rows if get_date(row) != "-"]
    return sorted(valid, key=parse_date)[-limit:]


def summarize(rows):
    result = []
    for kas in KASSEN:
        tmax = [get_tmax(row, kas) for row in rows]
        vdmax = [get_vdmax(row, kas) for row in rows]
        vdgem = [get_vdgem(row, kas) for row in rows]
        result.append({
            "kas": kas,
            "tmax_avg": avg(tmax),
            "tmax_max": max_or_none(tmax),
            "vdmax_avg": avg(vdmax),
            "vdmax_max": max_or_none(vdmax),
            "vdgem_avg": avg(vdgem),
        })
    return result


def timeseries(rows):
    series = []
    for row in latest_rows(rows, 20):
        ref_vd = get_vdmax(row, REFERENCE_KAS)
        ref_temp = get_tmax(row, REFERENCE_KAS)
        values = {}
        for kas in KASSEN:
            vdmax = get_vdmax(row, kas)
            tmax = get_tmax(row, kas)
            values[kas] = {
                "vdmax": vdmax,
                "tmax": tmax,
                "vd_delta": None if vdmax is None or ref_vd is None else vdmax - ref_vd,
                "temp_delta": None if tmax is None or ref_temp is None else tmax - ref_temp,
            }
        series.append({"date": get_date(row), "values": values})
    return series


def build_analysis():
    rows, source = load_rows()
    summary = summarize(rows)
    ref = next((item for item in summary if item["kas"] == REFERENCE_KAS), None)

    differences = []
    if ref:
        for item in summary:
            differences.append({
                "kas": item["kas"],
                "dtmax": None if item["tmax_avg"] is None or ref["tmax_avg"] is None else item["tmax_avg"] - ref["tmax_avg"],
                "dvdmax": None if item["vdmax_avg"] is None or ref["vdmax_avg"] is None else item["vdmax_avg"] - ref["vdmax_avg"],
                "dvdgem": None if item["vdgem_avg"] is None or ref["vdgem_avg"] is None else item["vdgem_avg"] - ref["vdgem_avg"],
            })

    critical = []
    for row in rows:
        values = [(kas, get_vdmax(row, kas)) for kas in KASSEN]
        values = [(kas, value) for kas, value in values if value is not None]
        if values:
            kas, value = max(values, key=lambda item: item[1])
            critical.append({"date": get_date(row), "kas": kas, "vd": value})
    critical = sorted(critical, key=lambda item: item["vd"], reverse=True)[:10]

    return {
        "status": "ok",
        "reference_kas": REFERENCE_KAS,
        "rows_count": len(rows),
        "source": source,
        "summary": summary,
        "differences": differences,
        "critical": critical,
        "timeseries": timeseries(rows),
        "debug": debug_info(),
    }


def axis_bounds(values):
    nums = [v for v in values if v is not None]
    if not nums:
        return 0, 1
    low = min(nums)
    high = max(nums)
    if low == high:
        return low - 1, high + 1
    pad = (high - low) * 0.12
    return low - pad, high + pad


def render_line_chart(title, series, metric, unit=""):
    width = 1080
    height = 430
    left = 64
    right = 28
    top = 34
    bottom = 108
    plot_w = width - left - right
    plot_h = height - top - bottom

    all_values = [
        item["values"].get(kas, {}).get(metric)
        for item in series
        for kas in KASSEN
    ]
    y_min, y_max = axis_bounds(all_values)

    def x_pos(index):
        if len(series) <= 1:
            return left + plot_w / 2
        return left + (index / (len(series) - 1)) * plot_w

    def y_pos(value):
        if value is None:
            return None
        return top + ((y_max - value) / (y_max - y_min)) * plot_h

    grid = []
    y_labels = []
    for i in range(5):
        value = y_min + (y_max - y_min) * i / 4
        y = y_pos(value)
        grid.append(f'<line x1="{left}" y1="{y:.1f}" x2="{width-right}" y2="{y:.1f}" stroke="#d8e0da" />')
        y_labels.append(f'<text x="{left-10}" y="{y+4:.1f}" text-anchor="end" font-size="12" fill="#52645b">{fmt(value)}</text>')

    day_labels = []
    for index, item in enumerate(series):
        x = x_pos(index)
        label = html.escape(day_label(item["date"]))
        day_labels.append(f'<line x1="{x:.1f}" y1="{height-bottom}" x2="{x:.1f}" y2="{height-bottom+6}" stroke="#9bad9f" />')
        day_labels.append(
            f'<text x="{x:.1f}" y="{height-34}" text-anchor="end" font-size="12" fill="#52645b" transform="rotate(-45 {x:.1f} {height-34})">{label}</text>'
        )

    lines = []
    points = []
    legend = []
    for kas in KASSEN:
        color = KAS_COLORS[kas]
        coords = []
        for index, item in enumerate(series):
            value = item["values"].get(kas, {}).get(metric)
            y = y_pos(value)
            if y is None:
                continue
            x = x_pos(index)
            coords.append(f"{x:.1f},{y:.1f}")
            points.append(
                f'<circle cx="{x:.1f}" cy="{y:.1f}" r="{3.6 if kas == REFERENCE_KAS else 2.7}" fill="{color}">'
                f'<title>Kas {kas} | {html.escape(item["date"])} | {fmt(value)} {html.escape(unit)}</title></circle>'
            )
        if len(coords) >= 2:
            width_line = 4 if kas == REFERENCE_KAS else 2.3
            lines.append(
                f'<polyline points="{" ".join(coords)}" fill="none" stroke="{color}" stroke-width="{width_line}" '
                f'stroke-linecap="round" stroke-linejoin="round" />'
            )
        legend.append(
            f'<span class="legend-item"><span class="legend-line" style="background:{color};"></span>'
            f'Kas {kas}{" referentie" if kas == REFERENCE_KAS else ""}</span>'
        )

    zero_line = ""
    if metric.endswith("_delta") and y_min < 0 < y_max:
        y0 = y_pos(0)
        zero_line = f'<line x1="{left}" y1="{y0:.1f}" x2="{width-right}" y2="{y0:.1f}" stroke="#222" stroke-dasharray="4,4" />'

    return f"""
    <div class="chart-card">
      <h2>{html.escape(title)}</h2>
      <div class="legend">{''.join(legend)}</div>
      <div class="svg-wrap">
        <svg viewBox="0 0 {width} {height}" role="img" aria-label="{html.escape(title)}">
          <rect x="0" y="0" width="{width}" height="{height}" rx="12" fill="#fff" />
          {''.join(grid)}
          {zero_line}
          <line x1="{left}" y1="{top}" x2="{left}" y2="{height-bottom}" stroke="#9bad9f" />
          <line x1="{left}" y1="{height-bottom}" x2="{width-right}" y2="{height-bottom}" stroke="#9bad9f" />
          {''.join(y_labels)}
          {''.join(day_labels)}
          {''.join(lines)}
          {''.join(points)}
        </svg>
      </div>
      <p class="chart-note">Alle dagen in deze grafiek staan onderaan als dag-maand. Laatste {len(series)} dagen.</p>
    </div>
    """


def td(value):
    return f"<td>{html.escape(str(value))}</td>"


def table(headers, rows):
    head = "".join(f"<th>{html.escape(str(h))}</th>" for h in headers)
    body = "".join("<tr>" + "".join(td(c) for c in row) + "</tr>" for row in rows)
    return f"<table><tr>{head}</tr>{body}</table>"


def render_page(data):
    warning = ""
    if data["rows_count"] == 0:
        debug = data.get("debug", {})
        warning = f"""
        <div class="warning">
            <b>Geen historische data gevonden.</b><br>
            Excel={debug.get('excel_exists')}, CSV-map={debug.get('csv_dir_exists')}, CSV-bestand={debug.get('csv_file_exists')}.<br>
            Remote fallback: {html.escape(str(debug.get('remote_csv_url')))}
        </div>
        """

    series = data.get("timeseries", [])
    charts = ""
    if series:
        charts = (
            render_line_chart("VD max per kas - laatste 20 dagen", series, "vdmax", "VD")
            + render_line_chart("Temperatuur max per kas - laatste 20 dagen", series, "tmax", "°C")
            + render_line_chart("Verschil VD max t.o.v. Kas4 - laatste 20 dagen", series, "vd_delta", "VD")
        )

    summary_rows = [
        [f"Kas {i['kas']}", f"{fmt(i['tmax_avg'])} °C", f"{fmt(i['tmax_max'])} °C", fmt(i["vdmax_avg"]), fmt(i["vdmax_max"]), fmt(i["vdgem_avg"])]
        for i in data["summary"]
    ]
    diff_rows = [
        [f"Kas {i['kas']}", f"{fmt(i['dtmax'])} °C", fmt(i["dvdmax"]), fmt(i["dvdgem"])]
        for i in data["differences"]
    ]
    critical_rows = [[i["date"], f"Kas {i['kas']}", fmt(i["vd"])] for i in data["critical"]]

    return f"""<!doctype html>
<html lang="nl">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Klimaat Dashboard Kas4</title>
  <style>
    body {{ margin:0; font-family:Arial,sans-serif; background:#eef3f0; color:#13251d; }}
    header {{ background:#123c2c; color:white; padding:28px 36px; }}
    main {{ max-width:1240px; margin:auto; padding:28px 36px; }}
    .cards {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(210px,1fr)); gap:16px; margin-bottom:26px; }}
    .card, .chart-card {{ background:white; border-radius:14px; padding:18px; box-shadow:0 8px 20px rgba(0,0,0,.08); }}
    .value {{ font-size:28px; font-weight:700; margin-top:8px; }}
    .ok {{ color:#16803c; }}
    .warning {{ background:#fff3cd; border:1px solid #e2c46d; padding:14px 16px; border-radius:12px; margin-bottom:18px; line-height:1.55; }}
    table {{ width:100%; border-collapse:collapse; background:white; border-radius:14px; overflow:hidden; margin:18px 0 32px; box-shadow:0 8px 20px rgba(0,0,0,.06); }}
    th, td {{ padding:10px 12px; border-bottom:1px solid #e3e8e4; text-align:left; }}
    th {{ background:#dfeae3; }}
    .chart-card {{ margin:22px 0 28px; }}
    .chart-card h2 {{ margin:0 0 10px; }}
    .svg-wrap {{ width:100%; overflow-x:auto; }}
    .svg-wrap svg {{ width:100%; min-width:880px; height:auto; display:block; }}
    .legend {{ display:flex; flex-wrap:wrap; gap:10px 16px; margin:8px 0 12px; color:#32473a; font-size:14px; }}
    .legend-item {{ display:inline-flex; align-items:center; gap:6px; }}
    .legend-line {{ display:inline-block; width:24px; height:4px; border-radius:999px; }}
    .chart-note {{ margin:8px 0 4px; color:#52645b; font-size:13px; }}
  </style>
</head>
<body>
  <header><h1>Klimaat Dashboard</h1><div>Historische analyse met <b>Kas4</b> als referentiekas</div></header>
  <main>
    {warning}
    <div class="cards">
      <div class="card"><div>Status</div><div class="value ok">Online</div></div>
      <div class="card"><div>Referentie</div><div class="value">Kas4</div></div>
      <div class="card"><div>Datapunten</div><div class="value">{data['rows_count']}</div></div>
      <div class="card"><div>Bron</div><div>{html.escape(data['source'])}</div></div>
    </div>
    {charts}
    <h2>Samenvatting per kas</h2>
    {table(["Kas", "Gem. Tmax", "Hoogste Tmax", "Gem. VD max", "Hoogste VD max", "Gem. VD gemiddeld"], summary_rows)}
    <h2>Afwijking t.o.v. Kas4</h2>
    {table(["Kas", "Δ gem. Tmax", "Δ gem. VD max", "Δ gem. VD gemiddeld"], diff_rows)}
    <h2>Top 10 kritische VD-pieken</h2>
    {table(["Datum", "Kas", "VD max"], critical_rows)}
  </main>
</body>
</html>"""


@app.get("/health")
def health():
    return {"status": "ok", "reference_kas": REFERENCE_KAS}


@app.get("/debug/files")
def debug_files():
    return JSONResponse(debug_info())


@app.get("/api/data")
def api_data():
    return JSONResponse(build_analysis())


@app.get("/", response_class=HTMLResponse)
def dashboard():
    try:
        return HTMLResponse(render_page(build_analysis()))
    except Exception as exc:
        return HTMLResponse(
            f"<h1>Klimaat Dashboard</h1><p>App draait, maar analyse gaf een fout:</p><pre>{html.escape(str(exc))}</pre>",
            status_code=200,
        )
